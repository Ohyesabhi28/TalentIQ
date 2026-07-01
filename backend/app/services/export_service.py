"""
Export Service.

Handles compiled exports (Excel, PDF, JSON) of candidate rankings and reports.
"""

from __future__ import annotations

import io
import json
from uuid import UUID

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors

from app.exceptions import AppError
from app.logging_config import get_logger
from app.store.memory import InMemoryStore

logger = get_logger(__name__)


class ExportService:
    def __init__(self, store: InMemoryStore):
        self._store = store

    def _determine_recommendation(self, score: float) -> str:
        if score >= 85:
            return "Strong Fit"
        elif score >= 70:
            return "Good Fit"
        elif score >= 50:
            return "Average Fit"
        else:
            return "Poor Fit"

    async def export_to_excel(self, job_id: UUID) -> bytes:
        logger.info("Generating Excel export for job_id=%s", job_id)
        
        ranking = await self._store.get_ranking(job_id)
        if not ranking or not ranking.candidates:
            raise AppError("RANKING_NOT_FOUND", f"No rankings found for job {job_id}.", 404)
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidate Rankings"
        
        # Grid lines visible
        ws.views.sheetView[0].showGridLines = True
        
        # Styles
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="4F518C", end_color="4F518C", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border_thin = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        headers = [
            "Rank", "Candidate Name", "Overall Score", "Skill Match Score", 
            "Semantic Match Score", "Experience Score", "Projects Score", 
            "Education Score", "Certifications Score", "Matched Skills", 
            "Missing Skills", "Recommendation", "Interview Focus"
        ]
        
        ws.row_dimensions[1].height = 28
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
            
        for row_num, cand in enumerate(ranking.candidates, 2):
            insight_rec = await self._store.get_recommendation(cand.candidate_id)
            
            rec = insight_rec.insight.hiring_recommendation if insight_rec else self._determine_recommendation(cand.overall_score)
            focus = ", ".join(insight_rec.insight.interview_focus_areas) if insight_rec else ""
            
            b = cand.breakdown
            matched_skills = ", ".join(b.skill_match.matched_skills)
            missing_skills = ", ".join(b.skill_match.missing_skills)
            
            row_data = [
                row_num - 1,
                cand.candidate_name,
                cand.overall_score,
                b.skill_match.skill_score,
                b.semantic_similarity,
                b.experience_match.experience_score,
                b.project_relevance.project_score,
                b.education_match.education_score,
                b.certification_match.certification_score,
                matched_skills,
                missing_skills,
                rec,
                focus
            ]
            
            ws.row_dimensions[row_num].height = 20
            for col_num, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=val)
                cell.border = border_thin
                
                # Format scores as percentages/numbers
                if col_num in [1, 3, 4, 5, 6, 7, 8, 9]:
                    cell.alignment = align_center
                    if col_num != 1:
                        cell.number_format = '0.0'
                else:
                    cell.alignment = align_left
                    
        # Autofit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)
            
        file_stream = io.BytesIO()
        wb.save(file_stream)
        return file_stream.getvalue()

    async def export_to_pdf(self, job_id: UUID) -> bytes:
        logger.info("Generating PDF export for job_id=%s", job_id)
        
        ranking = await self._store.get_ranking(job_id)
        if not ranking or not ranking.candidates:
            raise AppError("RANKING_NOT_FOUND", f"No rankings found for job {job_id}.", 404)
            
        job_record = await self._store.get_job(job_id)
        job_title = "Technical Role"
        required_skills = []
        if job_record and job_record.job_description_id:
            jd = await self._store.get_jd_profile(job_record.job_description_id)
            if jd:
                job_title = jd.job_title
                required_skills = jd.required_skills
                
        # ReportLab setup
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        # Color Palette
        primary_color = colors.HexColor("#4F518C")
        text_color = colors.HexColor("#2C2C2C")
        
        # Paragraph Styles
        style_title = ParagraphStyle(
            name="TitleStyle",
            fontName="Helvetica-Bold",
            fontSize=26,
            leading=30,
            textColor=primary_color,
            alignment=1, # Center
            spaceAfter=20
        )
        style_subtitle = ParagraphStyle(
            name="SubTitleStyle",
            fontName="Helvetica",
            fontSize=14,
            leading=18,
            textColor=colors.HexColor("#555555"),
            alignment=1,
            spaceAfter=50
        )
        style_heading = ParagraphStyle(
            name="HeadingStyle",
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            textColor=primary_color,
            spaceBefore=15,
            spaceAfter=10,
            keepWithNext=True
        )
        style_body = ParagraphStyle(
            name="BodyStyle",
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=text_color,
            spaceAfter=8
        )
        style_body_bold = ParagraphStyle(
            name="BodyBoldStyle",
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=14,
            textColor=text_color,
            spaceAfter=8
        )
        
        story = []
        
        # ── COVER PAGE ──
        story.append(Spacer(1, 100))
        story.append(Paragraph("TalentIQ", ParagraphStyle(name="Brand", fontName="Helvetica-Bold", fontSize=32, leading=36, textColor=primary_color, alignment=1, spaceAfter=10)))
        story.append(Paragraph("Enterprise Sourcing &amp; Rank Dossier", style_title))
        story.append(Paragraph(f"Target Role: {job_title}", style_subtitle))
        story.append(Spacer(1, 100))
        story.append(Paragraph("Confidential Executive Report", ParagraphStyle(name="Conf", fontName="Helvetica-Oblique", fontSize=11, alignment=1, textColor=colors.HexColor("#999999"))))
        story.append(PageBreak())
        
        # ── EXECUTIVE SUMMARY ──
        story.append(Paragraph("Executive Summary", style_heading))
        story.append(Paragraph(
            f"This dossier summarizes the semantic analysis and scoring breakdown of candidate profiles evaluated for the "
            f"<b>{job_title}</b> position. A total of <b>{len(ranking.candidates)} candidates</b> were evaluated against the core "
            f"role specification and requirements.",
            style_body
        ))
        
        if required_skills:
            story.append(Paragraph(f"<b>Core Skills Scrutinized:</b> {', '.join(required_skills)}", style_body))
            
        story.append(Spacer(1, 15))
        
        # ── CANDIDATE RANKINGS TABLE ──
        story.append(Paragraph("Candidate Rankings &amp; Scorecards", style_heading))
        
        table_data = [["Rank", "Candidate Name", "Overall Score", "Fit Recommendation"]]
        for idx, cand in enumerate(ranking.candidates, 1):
            rec = self._determine_recommendation(cand.overall_score)
            table_data.append([idx, cand.candidate_name, f"{cand.overall_score}%", rec])
            
        t = Table(table_data, colWidths=[40, 200, 100, 180])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), primary_color),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (1,1), (1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#F9F9FB")),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E2E6")),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F5F5FA")]),
        ]))
        story.append(t)
        
        story.append(PageBreak())
        
        # ── DETAILED RECOMMENDATIONS & FOCUS AREAS ──
        story.append(Paragraph("Hiring Recommendations &amp; Gaps", style_heading))
        for cand in ranking.candidates[:3]: # top 3 detailed focus
            insight_rec = await self._store.get_recommendation(cand.candidate_id)
            if insight_rec:
                insight = insight_rec.insight
                story.append(Paragraph(f"<b>Candidate: {cand.candidate_name} ({cand.overall_score}%)</b>", style_body_bold))
                story.append(Paragraph(f"<b>Summary:</b> {insight.overall_summary}", style_body))
                story.append(Paragraph(f"<b>Top Strengths:</b> {', '.join(insight.top_strengths)}", style_body))
                story.append(Paragraph(f"<b>Top Weaknesses:</b> {', '.join(insight.top_weaknesses)}", style_body))
                story.append(Paragraph(f"<b>Interview Focus:</b> {', '.join(insight.interview_focus_areas)}", style_body))
                story.append(Spacer(1, 10))
                
        # Footer build
        def add_footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.HexColor("#777777"))
            canvas.drawString(36, 20, "TalentIQ Platform Sourcing Report")
            canvas.drawRightString(letter[0]-36, 20, f"Page {doc.page}")
            canvas.restoreState()
            
        doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
        return buffer.getvalue()

    async def export_to_json(self, job_id: UUID) -> dict:
        logger.info("Generating JSON export for job_id=%s", job_id)
        
        ranking = await self._store.get_ranking(job_id)
        if not ranking:
            raise AppError("RANKING_NOT_FOUND", f"No ranking found for job {job_id}.", 404)
            
        job_record = await self._store.get_job(job_id)
        
        # Pull all insights
        insights_list = []
        for cand in ranking.candidates:
            insight_rec = await self._store.get_recommendation(cand.candidate_id)
            if insight_rec:
                insights_list.append(insight_rec.insight.model_dump())
                
        return {
            "job_id": str(job_id),
            "status": job_record.status.value if job_record else None,
            "ranking": ranking.model_dump(),
            "insights": insights_list
        }
